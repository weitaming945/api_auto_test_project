# -*- coding:utf-8 -*-
# @Time:2021/8/2 14:22
# @Author:lucky
# @Email:
# @File:handle_excel.py


import openpyxl

class HandleExcel:
    """
    读 写 excel
    """

    def __init__(self,filename,sheetname=None):
        """
        读写都有的参数
        :param filename: 文件路径
        :param sheetname: 表单名
        """
        self.filename = filename
        self.sheetname = sheetname

    def read_excel(self):

        # 加载excel文件为工作薄对象
        workbook = openpyxl.load_workbook(self.filename)
        # 选中表单
        sheet = workbook[self.sheetname]
        # 获取表单中所有行
        res = list(sheet.rows)
        #print(res) # [(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>
        # 获取表单中第一行的值，返回list
        title = [ i.value for i in res[0]]
        # print(title)
        #建一个空list，用来保存第一行与其他行组合的字典值
        cases = []
        # 取除第一行外的其他行的值
        for item in res[1:]:
            # 取每一行的值，返回list
            data = [i.value for i in item]
            # 使用zip函数，将第一行与其他行组合成字典
            dic = dict(zip(title,data))
            # 每循环一次，将字典添加到cases列表中
            cases.append(dic)

        #对数据处理,去掉\n,需要加上\将\n转义
        # print(str(cases).replace('\\n',''))
        cases=eval(str(cases).replace('\\n',''))

        return cases



    def read_excel_api_cases(self):
        """
        读取所有sheet的数据，保存为data
        不执行的sheet以#号开头
        :return:
        """
        # 加载excel文件为工作薄对象
        workbook = openpyxl.load_workbook(self.filename)
        #所有的sheet，返回为列表
        # sheets=workbook.get_sheet_names()
        sheets=workbook.sheetnames
        # print(sheets)
        #定义一个空列表，用于接收从excel中取的数据
        cases_api =[]

        for sheet in sheets:
            # print(sheet,type(sheet))
            #判断是否以API开头，#开头的排除
            if sheet.startswith('API') or sheet.startswith("api"):
                self.sheetname=sheet
                # 选中表单
                sheet = workbook[self.sheetname]
                # 获取表单中所有行
                res = list(sheet.rows)
                # print(res) # [(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>
                # 获取表单中第一行的值，返回list
                title = [i.value for i in res[0]]
                # print(title)
                # 建一个空list，用来保存第一行与其他行组合的字典值
                # 取除第一行外的其他行的值
                for item in res[1:]:
                    # 取每一行的值，返回list
                    data = [i.value for i in item]
                    # 使用zip函数，将第一行与其他行组合成字典
                    dic = dict(zip(title, data))
                    # 每循环一次，将字典添加到cases列表中
                    cases_api.append(dic)
        return cases_api

    def read_excel_web_cases(self):
        """
        读取所有sheet的数据，保存为data
        不执行的sheet以#号开头
        :return:
        """
        # 加载excel文件为工作薄对象
        workbook = openpyxl.load_workbook(self.filename)
        #所有的sheet，返回为列表get_sheet_names()方法被弃用
        # sheets=workbook.get_sheet_names()
        sheets=workbook.sheetnames
        # print(sheets)
        #定义一个空列表，用于接收从excel中取的数据
        cases_web =[]

        for sheet in sheets:
            # print(sheet,type(sheet))
            #判断是否以#开头，#开头的排除
            if sheet.startswith('WEB') or sheet.startswith('web'):
                pass
            else:
                self.sheetname=sheet
                # 选中表单
                sheet = workbook[self.sheetname]
                # 获取表单中所有行
                res = list(sheet.rows)
                # print(res) # [(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>
                # 获取表单中第一行的值，返回list
                title = [i.value for i in res[0]]
                # print(title)
                # 建一个空list，用来保存第一行与其他行组合的字典值
                # 取除第一行外的其他行的值
                for item in res[1:]:
                    # 取每一行的值，返回list
                    data = [i.value for i in item]
                    # 使用zip函数，将第一行与其他行组合成字典
                    dic = dict(zip(title, data))
                    # 每循环一次，将字典添加到cases列表中
                    cases_web.append(dic)
        return cases_web





    def write_excel(self,row,column,value):
        """
        写入数据
        :param row: 行
        :param column:列
        :param value: 值
        :return:
        """
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook[self.sheetname]
        sheet.cell(row=row,column=column,value=value)
        workbook.save(self.filename)





if __name__ == '__main__':

    from common.handle_excel import HandleExcel
    import os
    import unittest
    from common.handle_path import data_dir

    # 读excel数据
    test_data = HandleExcel(os.path.join(data_dir, 'psm.xlsx'))
    # 生成测试数据
    data = test_data.read_excel_api_cases()
    # print(data)
    print(test_data)


